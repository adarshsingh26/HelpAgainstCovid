import openpyxl
import requests

LOCATION_IQ_API_KEY = "pk.0db4987d09b5a2c6b19a15ecd68e002b"
file = "cat1.xlsx"
wb = openpyxl.load_workbook(file, read_only=True)
ws = wb.active


def get_data(district, state, country):
    data = dict()
    id = 0
    district = district.upper().strip()
    for row in ws.rows:
        for cell in row:
            if cell.value.replace("\n", "") == district:
                hospital = ws.cell(row=cell.row, column=3).value.replace("\n", "")
                beds = ws.cell(row=cell.row, column=11).value
                query_data = hospital + " " + district + " " + state
                query = "%20".join(query_data.split(" "))
                r = requests.get("https://us1.locationiq.com/v1/search.php?key=" + LOCATION_IQ_API_KEY + "&q=" + query + "&format=json")
                if isinstance(r.json(), dict):
                    print("AP1:Unable to geocode the hospital " + hospital)
                    break
                response = r.json()[0]
                # print("Response received: ", response)
                if response["display_name"].upper() != (state + ", " + country).upper():
                    data[id] = {"hospital": hospital, "state": state, "district": district,
                                "address": response["display_name"], "latitude": response["lat"],
                                "longitude": response["lon"], "country": country, "beds": beds}
                    id = id + 1
                else:
                    print("API: " + hospital + " data not available")
        # if id == 4:   #temp
        #     break
    return data